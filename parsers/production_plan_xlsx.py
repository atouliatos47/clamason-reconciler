"""Weekly Production Plan import — same 'WK30'-style workbook Matt
already maintains (WK_30D_2026.xlsx, WK_35D_2026.xlsx, ...). The sheet
we need gets renamed every week (WK30, WK31, ..., WK35, ...), so
instead of hardcoding a sheet name, this looks for the one sheet whose
name starts with 'WK' followed by digits — distinct from the
workbook's other planning sheets ('4 wk Plan', '1 wk plan'), which
never match that pattern.

Both numbers come straight from two fixed cells on that sheet, not
computed here:
  - Plan Quantity: E1
  - Plan Hours:    F1

These are already the workbook's own totals (checked against a real
WK35 export: E1 is NOT simply the sum of the sheet's own 'Planned'
column further down — it's built some other way inside the workbook,
likely against a filtered or pivoted range this parser never sees).
Reading the two finished cells is deliberately simpler and more
reliable than trying to reverse-engineer whatever produces them, and
matches what was actually asked for: read E1 and F1, don't recompute.
"""
import re

from openpyxl import load_workbook

SHEET_NAME_PATTERN = re.compile(r'^WK\d+', re.IGNORECASE)


def _find_plan_sheet(wb):
    """Returns the first sheet name starting 'WK<digits>' (e.g. 'WK35'),
    or None if no such sheet exists in this workbook."""
    for sheet_name in wb.sheetnames:
        if SHEET_NAME_PATTERN.match(sheet_name.strip()):
            return sheet_name
    return None


def parse_production_plan(path):
    """Returns {'sheet_name', 'plan_quantity', 'plan_hours'}.
    Raises ValueError with a clear message if no matching sheet is
    found, or if E1/F1 aren't populated numbers — better to fail
    loudly here than silently save a zero."""
    wb = load_workbook(path, data_only=True, read_only=True)
    sheet_name = _find_plan_sheet(wb)
    if sheet_name is None:
        raise ValueError(
            "Could not find a sheet named like 'WK35' in this workbook. "
            "Expected the same layout as the WK30-style production plan "
            "sheet — check the file is the right export."
        )

    ws = wb[sheet_name]
    plan_quantity = ws['E1'].value
    plan_hours = ws['F1'].value

    if not isinstance(plan_quantity, (int, float)) or not isinstance(plan_hours, (int, float)):
        raise ValueError(
            f"Found sheet '{sheet_name}', but E1 and/or F1 aren't populated "
            "numbers. Check the file actually has this week's plan filled in."
        )

    return {
        'sheet_name': sheet_name,
        'plan_quantity': round(float(plan_quantity), 2),
        'plan_hours': round(float(plan_hours), 2),
    }
